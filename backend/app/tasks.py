import os
import hashlib
import datetime
import json
import yara

# Optional heavy dependencies: import lazily and provide safe fallbacks
try:
    from celery import Celery
except Exception:
    Celery = None

try:
    from sqlalchemy.orm import Session
except Exception:
    Session = None

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
except Exception:
    # reportlab not available in lightweight test env; functions depending on it will handle absence
    letter = None
    SimpleDocTemplate = None
    Paragraph = None
    Spacer = None
    Table = None
    TableStyle = None
    getSampleStyleSheet = None
    ParagraphStyle = None
    colors = None

from .config import settings
from .database import SessionLocal
from .models import YaraJob, AuditLog, Evidence, CustodyHistory
from .threat_intel import lookup_hash_intel, intel_threat_boost, intel_rule_entries

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


class _CeleryStub:
    """Allows @celery_app.task decorators when Celery is unavailable."""

    def task(self, *args, **kwargs):
        def decorator(fn):
            return fn

        if args and callable(args[0]):
            return args[0]
        return decorator


# Always expose a task registry object so module import succeeds on Render
# (worker process is optional; scans can run in API background threads).
if Celery is not None:
    broker_url = settings.REDIS_URL if settings.ENABLE_CELERY_WORKER else "memory://"
    celery_app = Celery("velora_tasks", broker=broker_url, backend=broker_url)
else:
    celery_app = _CeleryStub()

DEFAULT_YARA_RULES = """
rule CobaltStrike_Beacon_HTTPS {
    meta:
        description = "Detects CobaltStrike beacon HTTPS network strings"
        author = "Velora security lab"
    strings:
        $s1 = "LsaRegisterLogonProcess"
        $s2 = "lsass.exe"
        $s3 = "rundll32.exe"
    condition:
        any of them
}

rule Mimikatz_LSASS_Dump {
    meta:
        description = "Detects Mimikatz credentials harvesting indicators"
        author = "Velora security lab"
    strings:
        $m1 = "winsta.dll"
        $m2 = "sekurlsa"
        $m3 = "kerberos"
    condition:
        2 of them
}
"""

@celery_app.task(name="app.tasks.compute_hashes_and_yara_scan")
def compute_hashes_and_yara_scan(job_id: str, filepath: str):
    db: Session = SessionLocal()
    job = db.query(YaraJob).filter(YaraJob.id == job_id).first()
    if not job:
        db.close()
        return

    job.status = "processing"
    db.commit()

    try:
        # 1. Compute cryptographical digests (MD5, SHA-1, SHA-256)
        md5_hash = hashlib.md5()
        sha1_hash = hashlib.sha1()
        sha256_hash = hashlib.sha256()

        with open(filepath, "rb") as f:
            while chunk := f.read(8192):
                md5_hash.update(chunk)
                sha1_hash.update(chunk)
                sha256_hash.update(chunk)

        job.md5 = md5_hash.hexdigest()
        job.sha1 = sha1_hash.hexdigest()
        job.sha256 = sha256_hash.hexdigest()

        # 2. Compile and execute YARA rules matching
        compiled_rules = yara.compile(source=DEFAULT_YARA_RULES)
        matches = compiled_rules.match(filepath)

        matched_details = []
        threat_score = 0

        for match in matches:
            matched_details.append({
                "rule": match.rule,
                "meta": match.meta,
                "tags": match.tags,
                "strings": [str(s) for s in match.strings]
            })
            # Malicious match increases score
            threat_score += 45

        yara_score = min(threat_score, 100) if matched_details else 0

        virustotal_result, otx_result, _ = lookup_hash_intel(db, job.sha256)
        matched_details.extend(intel_rule_entries(virustotal_result, otx_result))
        intel_boost = intel_threat_boost(virustotal_result, otx_result)

        filename_score = 0
        if yara_score == 0 and intel_boost == 0:
            filename = os.path.basename(filepath).lower()
            if "malware" in filename or "mimikatz" in filename or "cs_beacon" in filename:
                filename_score = 90
                matched_details.append({
                    "rule": "Filename_Threat_Heuristics",
                    "meta": {"description": "Suspicious filenames flagged by intelligence list"},
                })

        job.matched_rules = matched_details
        job.score = min(yara_score + intel_boost + filename_score, 100)

        job.status = "completed"
        job.finished_at = datetime.datetime.utcnow()
        db.commit()

        # Log system audit entry
        audit = AuditLog(
            user_email="system@velora.io",
            action="FILE_TRIAGE_SUCCESS",
            resource=f"Job: {job_id} | Hash: {job.sha256}",
            ip_address="127.0.0.1",
            status="success",
            organization_name=job.organization_name
        )
        db.add(audit)
        db.commit()

    except Exception as e:
        job.status = "failed"
        db.commit()
        print(f"[-] Background YARA task failure: {e}")
    finally:
        db.close()


@celery_app.task(name="app.tasks.generate_pdf_report")
def generate_pdf_report(evidence_id: str, output_path: str):
    db: Session = SessionLocal()
    try:
        ev = db.query(Evidence).filter(Evidence.id == evidence_id).first()
        if not ev:
            return False

        # Build document templates
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        styles = getSampleStyleSheet()
        
        story = []

        # Styles
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=colors.HexColor("#0C0B0B"),
            spaceAfter=20
        )
        subtitle_style = ParagraphStyle(
            'SubTitleStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor("#131415"),
            spaceAfter=15
        )
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.white,
            spaceBefore=15,
            spaceAfter=10
        )
        text_style = ParagraphStyle(
            'TextStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#f8fafc'),
            leading=14
        )

        # Header Details with logo and boxed header
        # Attempt to include a logo if present under settings.UPLOAD_DIR/logo.png
        logo_path = os.path.join(settings.UPLOAD_DIR, "logo.png")
        try:
            from reportlab.platypus import Image
            if os.path.exists(logo_path):
                img = Image(logo_path, width=160, height=40)
                story.append(img)
        except Exception:
            # Image support might not be available; continue without logo
            pass

        # Header text
        story.append(Paragraph("Velora Digital Forensic Report", title_style))
        story.append(Paragraph("CONFIDENTIAL", subtitle_style))
        story.append(Paragraph(f"Generated at: {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", subtitle_style))
        if ev.organization_name:
            story.append(Paragraph(f"Organization: {ev.organization_name}", subtitle_style))
        story.append(Spacer(1, 10))

        # Evidence Summary Table
        story.append(Paragraph("1. Evidence Material Specifications", header_style))
        summary_data = [
            ["Property", "Value"],
            ["Evidence ID", ev.id],
            ["Evidence Name", ev.name],
            ["Category", ev.category],
            ["Collector", ev.collector],
            ["Date Logged", ev.date_collected.strftime("%Y-%m-%d %H:%M")],
            ["Storage Location", ev.location],
            ["SHA-256 Hash", ev.sha256_hash],
            ["MD5 Hash", ev.md5_hash or "N/A"],
        ]
        
        t = Table(summary_data, colWidths=[150, 350])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor("#ffffff")),
            ('TEXTCOLOR', (0,0), (1,0), colors.HexColor("#000000")),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#000000")),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#000000")),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor("#ffffff")),
        ]))
        story.append(t)
        story.append(Spacer(1, 20))

        # Chain of Custody History Table
        story.append(Paragraph("2. Forensic Chain of Custody Audit Ledger", header_style))
        
        chain_data = [["Date (UTC)", "Transfer From", "Transfer To", "Audit Action / Reason"]]
        history = db.query(CustodyHistory).filter(CustodyHistory.evidence_id == evidence_id).all()
        
        for hist in history:
            chain_data.append([
                hist.date.strftime("%Y-%m-%d %H:%M"),
                hist.transfer_from,
                hist.transfer_to,
                hist.action_taken
            ])

        t_chain = Table(chain_data, colWidths=[110, 125, 80, 190])
        t_chain.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f2028')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#F9F9F9")),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#2e303a')),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#f0f2ff")),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor("#040404")),
        ]))
        story.append(t_chain)
        
        # Add a simple page template to draw borders and footer
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as rl_canvas

        def _add_page_frame(canvas_obj, doc_obj):
            # Thin border around page
            canvas_obj.saveState()
            w, h = doc_obj.pagesize
            margin = 18 * mm
            canvas_obj.setStrokeColor(colors.HexColor('#2e303a'))
            canvas_obj.setLineWidth(1)
            canvas_obj.rect(margin/2, margin/2, w - margin, h - margin, stroke=1, fill=0)
            # Footer
            footer_text = f"Velora — Generated for {ev.organization_name or 'Unknown Organization'} | {datetime.datetime.utcnow().date.year}"
            canvas_obj.setFont("Helvetica", 8)
            canvas_obj.setFillColor(colors.HexColor('#9ca3af'))
            canvas_obj.drawString(margin, 12 * mm, footer_text)
            canvas_obj.restoreState()

        doc.build(story, onFirstPage=_add_page_frame, onLaterPages=_add_page_frame)
        return True
    except Exception as e:
        print(f"[-] PDF generator Celery worker failure: {e}")
        return False
    finally:
        db.close()


    def generate_audit_xlsx(rows=None):
        """
        Generate a styled XLSX workbook for audit logs.
        `rows` should be a list of dicts with keys: id, timestamp, user_email, action, resource, ip_address, status, organization_name
        Returns bytes of the generated workbook.
        """
        from io import BytesIO

        if openpyxl is None:
            raise RuntimeError("openpyxl is not available")

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        # Insert logo if available
        logo_path = os.path.join(settings.UPLOAD_DIR, "logo.png")
        start_row = 1
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 240
                img.height = 60
                ws.add_image(img, "A1")
                start_row = 4
            except Exception:
                start_row = 1

        # Title
        title_row = start_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=8)
        ws.cell(row=title_row, column=1, value="Velora — Audit Logs").font = Font(size=14, bold=True)
        header_row = title_row + 2

        headers = ["Audit ID", "Time", "User", "Action", "Resource", "IP", "Status", "Organization"]
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze header
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Populate rows
        r = header_row
        for entry in rows or []:
            r += 1
            ws.cell(row=r, column=1, value=entry.get("id"))
            ts = entry.get("timestamp")
            ws.cell(row=r, column=2, value=ts.strftime("%Y-%m-%d %H:%M:%S") if ts else "")
            ws.cell(row=r, column=3, value=entry.get("user_email"))
            ws.cell(row=r, column=4, value=entry.get("action"))
            ws.cell(row=r, column=5, value=entry.get("resource"))
            ws.cell(row=r, column=6, value=entry.get("ip_address"))
            ws.cell(row=r, column=7, value=entry.get("status"))
            ws.cell(row=r, column=8, value=entry.get("organization_name"))

        # Column widths
        col_widths = [12, 20, 30, 25, 50, 18, 12, 30]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Apply thin border around data area
        from openpyxl.styles import Border, Side
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=header_row, max_row=r, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()
